import json
import requests
from datetime import datetime

from openpyxl import Workbook
file = open("templates/temp_logs.html","w+")

content = " <!DOCTYPE html><html><head>"
file.write(content)

def process_state(users,response,header,choices):
	content = [] 
	for user in users:
		t=user["profile"]
		l=[]
		for key in choices:
			try:
				p1=t[key]
				l.append(p1)

			except KeyError as e:
				l.append(" ")
		content.append(l)

	         
	if(response.headers["link"]):
		link_headers = response.headers["link"] # api call
		link_headers = list(link_headers.split(","))
		if (len(link_headers) >1):
			self_link = link_headers[0].split(";")[0].strip(" <>")
			next_link = str(link_headers[1].split(";")[0]).strip(" <>")

			count=0

			while(len(link_headers)>1 and next_link!=""):
				final_url = next_link
				response = requests.get(final_url, headers=header)
				users = json.loads(response.text)
				for user in users:
					t=user["profile"]
					l=[]
					for key in choices:
						try:
							p1=t[key]
							l.append(p1)

						except KeyError as e:
							l.append(" ")
					content.append(l)				
				link_headers = response.headers["link"]
				link_headers = list(link_headers.split(","))

				self_link = link_headers[0].split(";")[0].strip(" <>")
				if(len(link_headers)>1):
					next_link = str(link_headers[1].split(";")[0]).strip(" <>")
				count+=1
				print(count,response,next_link)
				logs = "<h3>"+str(count)+" "+str(response)+" "+next_link+"</h3>"
				file.write(logs)

	return content



def write_to_file(content_staged,content_active,content,file) : 
   
    column_list= ["A","B","C","D","E","F","G","H","I","J","K","L","M"]
    column_list=column_list[:len(content)]
    columns = content
    book = Workbook()
    ws1 = book.active
    ws1 = book.create_sheet("Sheet_A",0)
    ws1.title = "Staged state users"

    ws2 = book.create_sheet("Sheet_B",1)
    ws2.title="Active state users"



    count = 1
    for i in range(0,len(column_list)):
    	ws1.column_dimensions[column_list[i]].width = 35
    	col = column_list[i]+str(count)
    	ws1[col] = columns[i]
    count = count+1

    for i in range(0,len(content_staged)):
    	temp=content_staged[i]
    	for i in range(0,len(column_list)):
    		col = column_list[i]+str(count)
    		ws1[col] = temp[i]
    	count = count+1

    count = 1
    for i in range(0,len(column_list)):
    	ws2.column_dimensions[column_list[i]].width = 35
    	col = column_list[i]+str(count)
    	ws2[col] = columns[i]
    count = count+1

    for i in range(0,len(content_active)):
    	temp=content_active[i]
    	for i in range(0,len(column_list)):
    		col = column_list[i]+str(count)
    		ws2[col] = temp[i]
    	count = count+1

   

    book.save(file)




def main(Okta_tenant, API_token,choices):

	content="<h3> PROCESSING USERS IN STAGED STATE .. </h3>"
	file.write(content)
	print("PROCESSING USERS IN STAGED STATE")
	token = API_token
	final_url_staged = Okta_tenant+"/api/v1/users?filter=status eq \"STAGED\""

	header = {
	    'authorization': 'SSWS '+token,
	    'content-type': 'application/json'

	    }

	response_staged = requests.get(final_url_staged, headers=header)
	users_staged = json.loads(response_staged.text)
	

	content_staged = process_state(users_staged,response_staged,header,choices)

	content="<h3> PROCESSING USERS IN ACTIVE STATE .. </h3>"
	file.write(content)
	print("PROCESSING USERS IN ACTIVE STATE")
	final_url_active = Okta_tenant+"/api/v1/users?filter=status eq \"ACTIVE\""
	response_active = requests.get(final_url_active, headers=header)
	users_active = json.loads(response_active.text)
	content_active = process_state(users_active,response_active,header,choices)

	content = "number of active users "+str(len(content_active))
	file.write("<h3"+content+"</h3")
	content = "number of staged users"+str(len(content_staged))
	file.write("<h3"+content+"</h3")
	
	print("number of active users",len(content_active))
	print("number of staged users",len(content_staged))
	

	
	now = datetime.now()
	day= now.strftime("%m-%d-%Y")
	time = now.strftime("%H-%M-%S")


	path = "user_analysis_report_"+day+"_"+time+".xlsx"
	write_to_file(content_staged,content_active,choices,path)
	temp = "written to "+path
	

	processing_state=False
	
	return path

content = "</head><body></body></html>"
file.write(content)



#if __name__ == '__main__':
#	main()



